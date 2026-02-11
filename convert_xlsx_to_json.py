#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel形式のかるたデータをJSON形式に変換するスクリプト
入力: 260111geminiOCR.xlsx
出力: public/cards.json
"""

import openpyxl
import json
import os

def convert_excel_to_json(excel_path, output_path):
    """
    ExcelファイルをJSON形式に変換
    
    Args:
        excel_path (str): 入力Excelファイルパス
        output_path (str): 出力JSONファイルパス
    """
    print(f"読み込み中: {excel_path}")
    
    # Excelファイルを読み込み
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    cards = []
    
    # 2行目から101行目まで処理（1行目はヘッダー）
    for row_idx in range(2, 102):  # 2から101まで（100件）
        level = ws.cell(row_idx, 1).value  # A列: 大ピンチレベル
        initial = ws.cell(row_idx, 2).value  # B列: 頭文字
        content = ws.cell(row_idx, 3).value  # C列: 内容
        
        card = {
            "id": row_idx - 1,  # 札番号（1-100）
            "level": level,
            "initial": initial,
            "content": content
        }
        
        cards.append(card)
    
    # 出力ディレクトリが存在しない場合は作成
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"ディレクトリ作成: {output_dir}")
    
    # JSONファイルに書き込み（UTF-8、整形あり）
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(cards, f, ensure_ascii=False, indent=2)
    
    print(f"変換完了: {len(cards)}件のデータを {output_path} に出力しました")
    
    # サンプルデータを表示
    print("\n=== サンプルデータ（最初の3件）===")
    for i in range(min(3, len(cards))):
        print(f"札{cards[i]['id']}: レベル{cards[i]['level']}, {cards[i]['initial']}, {cards[i]['content']}")

if __name__ == "__main__":
    # ファイルパス設定
    excel_path = "260111geminiOCR.xlsx"
    output_path = os.path.join("public", "cards.json")
    
    # 変換実行
    try:
        convert_excel_to_json(excel_path, output_path)
        print("\n✓ 変換処理が正常に完了しました")
    except Exception as e:
        print(f"\n✗ エラーが発生しました: {e}")
        raise
